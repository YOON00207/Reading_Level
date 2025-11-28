import pandas as pd
import re
from openpyxl import load_workbook
import os


# 1. 한 파일 내 모든 시트 읽고 병합
def preprocess_single_excel(file_path):
    print(f"\n=== 파일 처리 시작: {file_path} ===")

    sheet_names = get_sheet_names(file_path)
    all_dataframes = []

    for sheet in sheet_names:
        df_sheet = process_one_sheet(file_path, sheet)
        if df_sheet is not None:
            all_dataframes.append(df_sheet)

    if not all_dataframes:
        print("⚠ 처리할 시트 데이터 없음")
        return None

    combined = pd.concat(all_dataframes, ignore_index=True)
    print("\n=== 파일 병합 완료 ===")
    return combined


# 2. 시트 목록 가져오기
def get_sheet_names(file_path):
    wb = load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    print(f"시트 목록: {sheets}")
    return sheets


# 3. 시트 하나 처리
def process_one_sheet(file_path, sheet_name):
    print(f"\n▶ 시트 처리: {sheet_name}")

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # 빈 시트 스킵
    if df.empty or df.shape[1] == 0:
        print("  → 빈 시트, 스킵")
        return None

    school, grade, student = extract_student_info(df)
    header_row = find_header_row(df)

    if header_row is None:
        print("  → 컬럼행을 찾지 못함, 스킵")
        return None

    data_df = build_dataframe(df, header_row)
    if data_df is None:
        return None

    # 메타데이터 추가
    data_df["학교"] = school
    data_df["학년"] = grade
    data_df["이름"] = student
    data_df["시트명"] = sheet_name

    print(f"  → {len(data_df)}행 처리 완료")
    return data_df

# 4. 상단 정보(학교/학년/이름) 추출
def extract_student_info(df):
    school = grade = student = None
    search_area = df.iloc[:10, :10]

    for r in range(search_area.shape[0]):
        for c in range(search_area.shape[1]):
            val = str(search_area.iloc[r, c]).strip()

            if val == "학교":
                school = search_area.iloc[r, c+1]
            elif val == "학년":
                grade = clean_grade(search_area.iloc[r, c+1])
            elif val == "이름":
                student = search_area.iloc[r, c+1]

    print(f"  학교={school}, 학년={grade}, 이름={student}")
    return school, grade, student


def clean_grade(grade_value):
    if pd.isna(grade_value):
        return None
    match = re.search(r"(\d+)", str(grade_value))
    return match.group(1) if match else grade_value


# 5. 컬럼 행 찾기 ("년" "월")
def find_header_row(df):
    for i in range(df.shape[0]):
        row_list = df.iloc[i].astype(str).tolist()
        if ("년" in row_list) and ("월" in row_list):
            print(f"  컬럼 행: {i+1}행")
            return i
    return None


# 6. 헤더 적용 후 본 데이터만 추출
def build_dataframe(df, header_row):
    headers = df.iloc[header_row].tolist()
    data = df.iloc[header_row + 1:].copy()
    data = data.dropna(how="all")

    if data.empty:
        print("  → 본문 데이터 없음")
        return None

    data.columns = headers
    data = data.reset_index(drop=True)
    return data


# 7. 통합 데이터 전처리(열 정리 등)
def clean_combined_dataframe(combined_df):
    print("\n=== 데이터 정리 시작 ===")

    df = combined_df.copy()

    target_cols = [
        '년', '월', '날짜', 'Date', '순번', '코드번호', '책제목',
        '레벨 ', '저자', '시리즈', '구분', '학교', '학년', '이름'
    ]
    df = df[target_cols]

    df = df.dropna(subset=['책제목'])

    df.columns = df.columns.str.replace(" ", "", regex=False)

    numeric_cols = ['년', '월', '날짜', '순번', '학년']
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

    df["레벨"] = (
        df["레벨"]
        .astype(str)
        .str.extract(r"(\d+)", expand=False)
        .astype(float)
        .astype("Int64")
        .fillna(-1)
    )

    print("=== 데이터 정리 완료 ===")
    return df


# 8. Date 자동 생성
def fill_missing_date(df):
    print("\n=== Date 자동 생성 중 ===")

    def make_date(row):
        if pd.isna(row['Date']) or row['Date'] == "":
            try:
                return pd.to_datetime(f"{int(row['년'])}-{int(row['월'])}-{int(row['날짜'])}")
            except:
                return pd.NaT
        return row['Date']

    df['Date'] = df.apply(make_date, axis=1)
    print("→ Date 생성 완료")
    return df


# 9. 엑셀 저장
def save_to_excel(df, output_path):
    df.to_excel(output_path, index=False)
    print(f"\n=== 엑셀 저장 완료: {output_path} ===")



# 10. 전체 실행 파이프라인
def main(file_path, output_path=None):
    if output_path is None:
        output_folder = 'output'
        os.makedirs(output_folder, exist_ok = True)
        output_path = os.path.join(output_folder, "전처리파일.xlsx")

    combined_df = preprocess_single_excel(file_path)
    cleaned_df = clean_combined_dataframe(combined_df)
    cleaned_df = fill_missing_date(cleaned_df)
    save_to_excel(cleaned_df, output_path)
    return cleaned_df


#Streamlit용으로 잠시 지우기
# if __name__ == "__main__":
#     file_path = input("전처리할 파일 경로를 입력하세요: ")
#     main(file_path)

